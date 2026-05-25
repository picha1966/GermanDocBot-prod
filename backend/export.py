# -*- coding: utf-8 -*-
"""
Export Module - експорт замовлень у CSV/Excel
"""

import os
import csv
import json
from datetime import datetime, timedelta
from typing import List, Optional
from io import BytesIO, StringIO


class ExportManager:
    """
    Менеджер експорту даних для адміна
    """
    
    def __init__(self, output_dir: str = None):
        """
        Args:
            output_dir: Директорія для збереження файлів експорту
        """
        self.output_dir = output_dir or os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'exports'
        )
        os.makedirs(self.output_dir, exist_ok=True)
    
    def export_orders_csv(
        self,
        orders: List[dict],
        filename: str = None
    ) -> str:
        """
        Експортувати замовлення в CSV файл
        
        Args:
            orders: Список замовлень
            filename: Назва файлу (опціонально)
            
        Returns:
            Шлях до створеного файлу
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'orders_export_{timestamp}.csv'
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Визначаємо колонки
        columns = [
            'id', 'user_id', 'username', 'first_name', 'last_name',
            'doc_type', 'status', 'price', 'discount', 'final_price',
            'promo_code', 'stripe_session_id', 'paid_at',
            'created_at', 'updated_at'
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
            writer.writeheader()
            
            for order in orders:
                # Обробляємо дані
                row = {col: order.get(col, '') for col in columns}
                writer.writerow(row)
        
        return filepath
    
    def export_orders_excel(
        self,
        orders: List[dict],
        filename: str = None
    ) -> Optional[str]:
        """
        Експортувати замовлення в Excel файл
        
        Args:
            orders: Список замовлень
            filename: Назва файлу (опціонально)
            
        Returns:
            Шлях до створеного файлу або None якщо openpyxl недоступний
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return None
        
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'orders_export_{timestamp}.xlsx'
        
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Orders'
        
        # Заголовки
        headers = [
            ('ID', 8),
            ('User ID', 12),
            ('Username', 15),
            ('Ім\'я', 15),
            ('Прізвище', 15),
            ('Тип документа', 15),
            ('Статус', 12),
            ('Ціна €', 10),
            ('Знижка €', 10),
            ('Фінальна ціна €', 12),
            ('Промокод', 12),
            ('Stripe Session', 25),
            ('Оплачено', 18),
            ('Створено', 18),
        ]
        
        # Стилі
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Записуємо заголовки
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Записуємо дані
        status_colors = {
            'pending': 'FFF2CC',    # Жовтий
            'paid': 'C6EFCE',       # Зелений
            'processing': 'BDD7EE', # Блакитний
            'ready': 'C6EFCE',      # Зелений
            'sent': 'E2EFDA',       # Світло-зелений
            'downloaded': 'E2EFDA', # Світло-зелений
            'cancelled': 'FFC7CE', # Червоний
            'failed': 'FFC7CE',    # Червоний
        }
        
        for row_num, order in enumerate(orders, 2):
            data = [
                order.get('id', ''),
                order.get('user_id', ''),
                order.get('username', ''),
                order.get('first_name', ''),
                order.get('last_name', ''),
                order.get('doc_type', ''),
                order.get('status', ''),
                order.get('price', 0),
                order.get('discount', 0),
                order.get('final_price', 0),
                order.get('promo_code', ''),
                order.get('stripe_session_id', ''),
                order.get('paid_at', '')[:16] if order.get('paid_at') else '',
                order.get('created_at', '')[:16] if order.get('created_at') else '',
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                
                # Форматуємо числа
                if col in [8, 9, 10]:  # Ціни
                    cell.number_format = '#,##0.00'
                
                # Колір статусу
                if col == 7:
                    status = str(value).lower()
                    if status in status_colors:
                        cell.fill = PatternFill(
                            start_color=status_colors[status],
                            end_color=status_colors[status],
                            fill_type='solid'
                        )
        
        # Автофільтр
        ws.auto_filter.ref = ws.dimensions
        
        # Заморожуємо заголовок
        ws.freeze_panes = 'A2'
        
        wb.save(filepath)
        return filepath
    
    def export_orders_bytes(
        self,
        orders: List[dict],
        format: str = 'csv'
    ) -> Optional[bytes]:
        """
        Експортувати в байти (для відправки через Telegram)
        
        Args:
            orders: Список замовлень
            format: 'csv' або 'excel'
            
        Returns:
            Байти файлу
        """
        if format == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                format = 'csv'
        
        if format == 'csv':
            output = StringIO()
            columns = [
                'id', 'user_id', 'username', 'doc_type', 'status',
                'price', 'final_price', 'paid_at', 'created_at'
            ]
            
            writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
            writer.writeheader()
            
            for order in orders:
                row = {col: order.get(col, '') for col in columns}
                writer.writerow(row)
            
            return output.getvalue().encode('utf-8-sig')
        
        else:  # excel
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            
            headers = ['ID', 'User', 'Тип', 'Статус', 'Ціна', 'Оплата', 'Дата']
            ws.append(headers)
            
            for order in orders:
                ws.append([
                    order.get('id'),
                    order.get('user_id'),
                    order.get('doc_type'),
                    order.get('status'),
                    order.get('final_price'),
                    order.get('paid_at', '')[:16] if order.get('paid_at') else '',
                    order.get('created_at', '')[:16] if order.get('created_at') else '',
                ])
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
    
    def generate_report_summary(
        self,
        orders: List[dict],
        days: int = 7
    ) -> str:
        """
        Генерувати текстовий звіт
        
        Args:
            orders: Список замовлень
            days: Період (днів)
            
        Returns:
            Форматований текст звіту
        """
        total_orders = len(orders)
        
        # Статистика по статусах
        status_counts = {}
        for order in orders:
            status = order.get('status', 'unknown')
            status_counts[status] = status_counts.get(status, 0) + 1
        
        # Фінансова статистика
        total_revenue = sum(order.get('final_price', 0) or 0 for order in orders if order.get('status') in ['paid', 'sent', 'downloaded'])
        total_discounts = sum(order.get('discount', 0) or 0 for order in orders)
        
        # Статистика по типах документів
        doc_counts = {}
        for order in orders:
            doc_type = order.get('doc_type', 'unknown')
            doc_counts[doc_type] = doc_counts.get(doc_type, 0) + 1
        
        # Формуємо звіт
        report = f"""
📊 **ЗВІТ ЗА {days} ДНІВ**
{'═' * 30}

📋 **ЗАМОВЛЕННЯ:**
• Всього: {total_orders}
"""
        
        status_emoji = {
            'pending': '⏳',
            'paid': '✅',
            'processing': '🔄',
            'ready': '📄',
            'sent': '📤',
            'downloaded': '✔️',
            'cancelled': '❌',
            'failed': '⚠️'
        }
        
        for status, count in status_counts.items():
            emoji = status_emoji.get(status, '❓')
            report += f"• {emoji} {status}: {count}\n"
        
        report += f"""
💰 **ФІНАНСИ:**
• Дохід: {total_revenue:.2f}€
• Знижки: {total_discounts:.2f}€

📁 **ПО ТИПАХ:**
"""
        
        for doc_type, count in sorted(doc_counts.items(), key=lambda x: -x[1]):
            report += f"• {doc_type}: {count}\n"
        
        report += f"""
{'═' * 30}
📅 Згенеровано: {datetime.now().strftime('%d.%m.%Y %H:%M')}
"""
        
        return report


# Глобальний екземпляр
export_manager = ExportManager()
